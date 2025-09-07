import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Класс для генерации отчетов в различных форматах."""

    def __init__(self, calculator):
        """Инициализация генератора отчетов."""
        self.calculator = calculator
        self.summary = calculator.summary()
        logger.debug(
            'ReportGenerator инициализирован с данными: %s',
            self.summary
        )

    def generate_excel_report(self, filename: str):
        """Генерирует Excel-отчет с результатами расчета теплоизоляции."""
        logger.info('Начата генерация Excel-отчета: %s', filename)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Расчет теплоизоляции'

            header_font = Font(name='Arial', size=14, bold=True)
            title_font = Font(name='Arial', size=16, bold=True)
            regular_font = Font(name='Arial', size=11)

            header_fill = PatternFill(
                start_color='366092',
                end_color='366092',
                fill_type='solid'
            )
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws.merge_cells('A1:C1')
            ws['A1'] = 'РАСЧЕТ ТЕПЛОИЗОЛЯЦИИ ДЛЯ НФС'
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:C2')
            ws['A2'] = (
                f'Дата создания: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            )
            ws['A2'].font = regular_font
            ws['A2'].alignment = Alignment(horizontal='center')

            row = 4
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = 'ИНФОРМАЦИЯ О МАТЕРИАЛЕ'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = Font(
                name='Arial',
                size=14,
                bold=True,
                color='FFFFFF'
            )
            row += 1

            for label, value in [
                ('Код материала:', self.summary['SKU']),
                ('Наименование:', self.summary['Наименование материала'])
            ]:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
                ws[f'B{row}'].font = regular_font
                row += 1

            row += 1

            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = 'ИСХОДНЫЕ ДАННЫЕ'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = Font(
                name='Arial',
                size=14,
                bold=True,
                color='FFFFFF'
            )
            row += 1

            for label, value in [
                ('Площадь здания:', f'{self.summary["Площадь фасада"]} м²'),
                ('Высота здания:', f'{self.summary["Высота здания"]} м'),
                ('Количество углов:', f'{self.summary["Количесто внешних углов здания"]} шт.'),
                ('Периметр здания:', f'{self.summary["Периметр"]} м')
            ]:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
                ws[f'B{row}'].font = regular_font
                row += 1

            row += 1

            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = 'РЕЗУЛЬТАТЫ РАСЧЕТА'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = Font(
                name='Arial',
                size=14,
                bold=True,
                color='FFFFFF'
            )
            row += 1

            for label, value in [
                (
                    'Общая площадь утепления:',
                    f'{self.summary["Площадь теплоизоляции"]} м²'
                ),
                (
                    'Количество листов:',
                    f'{self.summary["Количество МВП (шт)"]} шт.'
                ),
                (
                    'Общий объем материала:',
                    f'{self.summary["Объем МВП"]} м³'
                ),
                (
                    'Количество крепежа:',
                    f'{self.summary["Количество крепежа"]} шт.'
                )
            ]:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
                ws[f'B{row}'].font = regular_font
                row += 1

            row += 1

            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = 'РЕКОМЕНДАЦИИ ПО КРЕПЕЖУ'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = Font(
                name='Arial',
                size=14,
                bold=True,
                color='FFFFFF'
            )
            row += 1

            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = self.summary['Длина крепежа']
            ws[f'A{row}'].font = regular_font

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20

            for row_cells in ws.iter_rows(
                min_row=1,
                max_row=row,
                min_col=1,
                max_col=3
            ):
                for cell in row_cells:
                    if cell.value:
                        cell.border = border

            abs_path = os.path.abspath(filename)
            folder = os.path.dirname(abs_path)
            if folder and not os.path.exists(folder):
                os.makedirs(folder, exist_ok=True)
                logger.info(f'Создана директория: {folder}')

            logger.info(
                f'Пытаемся сохранить файл по абсолютному пути: {abs_path}'
            )
            wb.save(abs_path)
            logger.info('Excel-отчет успешно сохранен: %s', abs_path)

        except Exception as e:
            logger.exception('Ошибка при генерации Excel-отчета: %s', e)
            raise

    def generate_pdf_report(self, filename: str):
        """Генерирует PDF-отчет с результатами расчета теплоизоляции."""
        logger.info('Начата генерация PDF-отчета: %s', filename)
        try:
            pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName='Arial',
                fontSize=18,
                spaceAfter=30,
                alignment=1
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName='Arial',
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )

            normal_style = styles['Normal'].clone('normal_dejavu')
            normal_style.fontName = 'Arial'

            story = []

            story.append(Paragraph(
                'РАСЧЕТ ТЕПЛОИЗОЛЯЦИИ ДЛЯ НФС',
                title_style
            ))
            story.append(Paragraph(
                f'Дата создания: {datetime.now().strftime("%d.%m.%Y %H:%M")}',
                normal_style
            ))
            story.append(Spacer(1, 20))

            def add_table(title, data):
                story.append(Paragraph(title, heading_style))
                table = Table(data, colWidths=[3 * inch, 3 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                story.append(Spacer(1, 20))

            add_table('ИНФОРМАЦИЯ О МАТЕРИАЛЕ', [
                ['Наименование:', self.summary['Наименование материала']]
            ])

            add_table('ИСХОДНЫЕ ДАННЫЕ', [
                ['Площадь здания:', f'{self.summary["Площадь фасада"]} м²'],
                ['Высота здания:', f'{self.summary["Высота здания"]} м'],
                ['Количество углов:', f'{self.summary["Количесто внешних углов здания"]} шт.'],
                ['Периметр здания:', f'{self.summary["Периметр"]} м']
            ])

            add_table('РЕЗУЛЬТАТЫ РАСЧЕТА', [
                [
                    'Общая площадь утепления:',
                    f'{self.summary["Площадь теплоизоляции"]} м²'
                ],
                [
                    'Количество листов:',
                    f'{self.summary["Количество МВП (шт)"]} шт.'
                ],
                [
                    'Общий объем материала:',
                    f'{self.summary["Объем МВП"]} м³'
                ],
                [
                    'Количество крепежа:',
                    f'{self.summary["Количество крепежа"]} шт.'
                ]
            ])

            story.append(Paragraph('РЕКОМЕНДАЦИИ ПО КРЕПЕЖУ', heading_style))
            story.append(Paragraph(
                self.summary['Длина крепежа'],
                normal_style
            ))

            doc.build(story)
            logger.info('PDF-отчет успешно сохранен: %s', filename)

        except Exception as e:
            logger.exception('Ошибка при генерации PDF-отчета: %s', e)
            raise
